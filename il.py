#------------------------------------------------------------------------------#
#            Copyright 2018 Rieckermann Engineering Operations                 #
#------------------------------------------------------------------------------#
# Description:                                                                 #
# This file creates a list of interlocks from a safety matrix.                 #
#------------------------------------------------------------------------------#
# Revision history:                                                            #
# Ver By               Date        CC        Note                              #
# 1   David Paspa      05-Apr-2018 NA        Initial design.                   #
#------------------------------------------------------------------------------#
# Import the python libraries:                                                 #
#------------------------------------------------------------------------------#
import argparse
from enum import Enum
import openpyxl
import os.path
import sys
from tqdm import trange
from time import sleep

import logging
logging.basicConfig(level=logging.ERROR)

#------------------------------------------------------------------------------#
# Declare the application title and calling arguments help:                    #
#------------------------------------------------------------------------------#
appTitle = 'Interlock Generator'
appVersion = '1'
parser = argparse.ArgumentParser(description='Generates a list of Interlocks from a Safety Matrix worksheet')
parser.add_argument('-c','--config', help='Path and file name of the input safety matrix worksheet file', required=True)
parser.add_argument('-s','--sheet', help='Name of the safety matrix worksheet', required=True)
parser.add_argument('-d','--delete', help='Delete old data and start new interlock list if "y"', required=True)
parser.add_argument('-a','--at', help='Replace at character with tag prefix number, such as 1, 2, 3 or 4', required=True)
parser.add_argument('-l','--dollar', help='Replace dollar character with equipment letter, such as M or S', required=True)
parser.add_argument('-n','--percent', help='Replace percent character with equipment prefix number, such as 1 or 2 for M1 or M2', required=True)
parser.add_argument('-p','--parent', help='Parent object to generate code files for', required=True)
args = vars(parser.parse_args())

#------------------------------------------------------------------------------#
# Declare global variables:                                                    #
#------------------------------------------------------------------------------#
connInstance = 0
idxInstance = -1
wb = ''

#------------------------------------------------------------------------------#
# Declare the error handling global variables and procedure:                   #
#------------------------------------------------------------------------------#
def errorHandler(errProc, eCode, *args):
    #--------------------------------------------------------------------------#
    # Get the application specific error message and output the error:         #
    #--------------------------------------------------------------------------#
    sMsg = errorMessage[eCode]

    #--------------------------------------------------------------------------#
    # Replace any error parameters:                                            #
    #--------------------------------------------------------------------------#
    i = 1
    for arg in args:
        sMsg = sMsg.replace('@' + str(i), str(arg))
        i = i + 1

    #--------------------------------------------------------------------------#
    # Close the progress bar:                                                  #
    #--------------------------------------------------------------------------#
    p.close()

    #--------------------------------------------------------------------------#
    # Output the error message and end:                                        #
    #--------------------------------------------------------------------------#
    print('\r\n')
    print(appTitle + ' Version ' + appVersion + '\r\n' + 'ERROR ' +
          str(eCode) + ' in Procedure ' + "'" + errProc + "'" + '\r\n' + '\r\n' + sMsg)
    print('\r\n')
    print(traceback.format_exception(*sys.exc_info()))
    sys.exit()

#------------------------------------------------------------------------------#
# Enumerate the error numbers and map the error messages:                      #
#------------------------------------------------------------------------------#
class errorCode(Enum):
    cannotCreateOutputSheet            = -1
    cannotOpenWorkbook                 = -2
    cannotReplace                      = -3
    fileNotExist                       = -4
    instanceNotFound                   = -5
    noMatrixWorksheet                  = -6
    noNamedRange                       = -7

errorMessage = {
    errorCode.cannotCreateOutputSheet  : 'Cannot create output worksheet @1',
    errorCode.cannotOpenWorkbook       : 'Cannot open workbook @1',
    errorCode.cannotReplace            : 'Cannot replace field @1 with value @2.',
    errorCode.fileNotExist             : 'Workbook file @1 does not exist.',
    errorCode.instanceNotFound         : 'Instance @1 does not exist in sheet tblInstance.',
    errorCode.noMatrixWorksheet        : 'Safety matrix worksheet @1 does not exist.',
    errorCode.noNamedRange             : 'Named range @1 does not exist.'
}

#------------------------------------------------------------------------------#
# Create a progress bar:                                                       #
#------------------------------------------------------------------------------#
p = trange(100, desc='Starting...', leave=False)

#------------------------------------------------------------------------------#
# Function main                                                                #
#                                                                              #
# Description:                                                                 #
# The main entry point for the program.                                        #
#------------------------------------------------------------------------------#
def main():
    #--------------------------------------------------------------------------#
    # Define the procedure name:                                               #
    #--------------------------------------------------------------------------#
    errProc = main.__name__

    #--------------------------------------------------------------------------#
    # Define global variables used in this procedure:                          #
    #--------------------------------------------------------------------------#
    global wb

    #--------------------------------------------------------------------------#
    # Get the interlock configuration workbook name and check it exists:       #
    #--------------------------------------------------------------------------#
    wbName = args['config']
    if not os.path.exists(wbName):
        errorHandler(errProc, errorCode.fileNotExist, wbName)

    #--------------------------------------------------------------------------#
    # Open the workbook:                                                       #
    #--------------------------------------------------------------------------#
    try:
        wb = openpyxl.load_workbook(wbName, data_only=True)
    except:
        errorHandler(errProc, errorCode.cannotOpenWorkbook, wbName)

    #--------------------------------------------------------------------------#
    # Check if the existing output worksheets are to be deleted:               #
    #--------------------------------------------------------------------------#
    deleteExisting = args['delete']
    if (deleteExisting.upper() == 'Y' or deleteExisting.upper() == 'YES'):
        #----------------------------------------------------------------------#
        # Delete the Critical Interlock output worksheet if exists:            #
        #----------------------------------------------------------------------#
        wsoName = 'tblInterlockCRIL'
        try:
            wso = wb[wsoName]
            wb.remove(wso)
        except:
            pass

        #----------------------------------------------------------------------#
        # Create a new blank Critical Interlock worksheet:                     #
        #----------------------------------------------------------------------#
        try:
            wb.create_sheet(wsoName)
            wso = wb[wsoName]
        except:
            errorHandler(errProc, errorCode.cannotCreateOutputSheet, wsoName)

        #----------------------------------------------------------------------#
        # Set the new Critical Interlock worksheet titles:                     #
        #----------------------------------------------------------------------#
        wso.cell(row=1, column=1).value = 'Instance'
        wso.cell(row=1, column=2).value = 'Class'
        wso.cell(row=1, column=3).value = 'IDX'
        wso.cell(row=1, column=4).value = 'Description'
        wso.cell(row=1, column=5).value = 'DescriptionIL'
        wso.cell(row=1, column=6).value = 'Interlock'
        wso.cell(row=1, column=7).value = 'Connection'

        #----------------------------------------------------------------------#
        # Delete the Non-Critical Interlock output worksheet if exists:        #
        #----------------------------------------------------------------------#
        wsoName = 'tblInterlockNCRIL'
        try:
            wso = wb[wsoName]
            wb.remove(wso)
        except:
            pass

        #----------------------------------------------------------------------#
        # Create a new blank Non-Critical Interlock worksheet:                 #
        #----------------------------------------------------------------------#
        try:
            wb.create_sheet(wsoName)
            wso = wb[wsoName]
        except:
            errorHandler(errProc, errorCode.cannotCreateOutputSheet, wsoName)

        #----------------------------------------------------------------------#
        # Set the new Non-Critical Interlock worksheet titles:                 #
        #----------------------------------------------------------------------#
        wso.cell(row=1, column=1).value = 'Instance'
        wso.cell(row=1, column=2).value = 'Class'
        wso.cell(row=1, column=3).value = 'IDX'
        wso.cell(row=1, column=4).value = 'Description'
        wso.cell(row=1, column=5).value = 'DescriptionIL'
        wso.cell(row=1, column=6).value = 'Interlock'
        wso.cell(row=1, column=7).value = 'Connection'

    #--------------------------------------------------------------------------#
    # Get the new output worksheet object references:                          #
    #--------------------------------------------------------------------------#
    wsoCR = wb['tblInterlockCRIL']
    wsoNCR = wb['tblInterlockNCRIL']

    #--------------------------------------------------------------------------#
    # Get the input safety matrix worksheet:                                   #
    #--------------------------------------------------------------------------#
    wsiName = args['sheet']
    try:
        wsi = wb[wsiName]
    except:
        errorHandler(errProc, errorCode.noMatrixWorksheet, wsiName)

    #--------------------------------------------------------------------------#
    # Get the parent object in the hierarchy:                                  #
    #--------------------------------------------------------------------------#
    sParent = args['parent']

    #--------------------------------------------------------------------------#
    # Get the placeholder replacement characters:                              #
    #--------------------------------------------------------------------------#
    at = args['at']
    dollar = args['dollar']
    percent = args['percent']

    #--------------------------------------------------------------------------#
    # Process the critical interlocks:                                         #
    #--------------------------------------------------------------------------#
    generateInterlocks(wb, wsi, wsoCR, wsiName, 'X', sParent, 50, at, dollar, percent)

    #--------------------------------------------------------------------------#
    # Process the non-critical interlocks:                                     #
    #--------------------------------------------------------------------------#
    generateInterlocks(wb, wsi, wsoNCR, wsiName, 'M', sParent, 50, at, dollar, percent)

    #--------------------------------------------------------------------------#
    # Save the changes if no error:                                            #
    #--------------------------------------------------------------------------#
    wb.save(wbName)

    #--------------------------------------------------------------------------#
    # Report completion regardless of error:                                   #
    #--------------------------------------------------------------------------#
    p.set_description('Interlock processing complete')
    p.refresh()
    p.close()

    #--------------------------------------------------------------------------#
    # Output a success message:                                                #
    #--------------------------------------------------------------------------#
    print('Congratulations... ' + wsiName + ' interlock code generation successful.')

#------------------------------------------------------------------------------#
# Function generateInterlocks                                                  #
#                                                                              #
# Description:                                                                 #
# Creates the interlock worksheet with the safety matrix as an interlock list. #
#------------------------------------------------------------------------------#
# Calling parameters:                                                          #
#                                                                              #
# wb                    The workbook object.                                   #
# wsi                   The input safety matrix worksheet object.              #
# wsiName               The output sheet name.                                 #
# sILMarker             The interlock marker, either "X" for critical          #
#                       interlocks or "M" for non-critical interlocks          #
#                       which are allowed to be manually overrridden.          #
# sParent               The parent object in the hierarchy.                    #
# pbwt                  The % weight of the procedure for the progress bar.    #
# at                    Replacement character for the at placeholder.          #
# dollar                Replacement character for the dollar placeholder.      #
# percent               Replacement character for the percent placeholder.     #
#------------------------------------------------------------------------------#
def generateInterlocks(wb, wsi, wso, wsiName, sILMarker, sParent, pbwt, at, dollar, percent):
    #--------------------------------------------------------------------------#
    # Define the procedure name:                                               #
    #--------------------------------------------------------------------------#
    errProc = generateInterlocks.__name__

    #--------------------------------------------------------------------------#
    # Use the global progress bar:                                             #
    #--------------------------------------------------------------------------#
    global connInstance
    global idxInstance
    global p

    #--------------------------------------------------------------------------#
    # Find the last row in the output sheet:                                   #
    #--------------------------------------------------------------------------#
    iRowOut = 2
    while (not wso.cell(row=iRowOut, column=1).value is None):
        iRowOut = iRowOut + 1

    #--------------------------------------------------------------------------#
    # Get the named ranges to orientate the data:                              #
    #--------------------------------------------------------------------------#
    try:
        sNamedRange = wsiName + 'End'
        colILEnd = wsi[list(wb.defined_names[sNamedRange].destinations)[0][1]].col_idx
        sNamedRange = wsiName + 'Name'
        colILName = wsi[list(wb.defined_names[sNamedRange].destinations)[0][1]].col_idx
        sNamedRange = wsiName + 'Code'
        colILCode = wsi[list(wb.defined_names[sNamedRange].destinations)[0][1]].col_idx
        sNamedRange = wsiName + 'Target'
        colILTarget = wsi[list(wb.defined_names[sNamedRange].destinations)[0][1]].col_idx
        rowILTarget = wsi[list(wb.defined_names[sNamedRange].destinations)[0][1]].row
        sNamedRange = wsiName + 'Source'
        colILSource = wsi[list(wb.defined_names[sNamedRange].destinations)[0][1]].col_idx
        rowILSource = wsi[list(wb.defined_names[sNamedRange].destinations)[0][1]].row
        sNamedRange = wsiName + 'State'
        colILState = wsi[list(wb.defined_names[sNamedRange].destinations)[0][1]].col_idx
        sNamedRange = wsiName + 'Function'
        colILFunction = wsi[list(wb.defined_names[sNamedRange].destinations)[0][1]].col_idx
        sNamedRange = wsiName + 'FunctionEnd'
        colILFunctionEnd = wsi[list(wb.defined_names[sNamedRange].destinations)[0][1]].col_idx
        sNamedRange = wsiName + 'Description'
        colILDescription = wsi[list(wb.defined_names[sNamedRange].destinations)[0][1]].col_idx
        sNamedRange = wsiName + 'TargetClass'
        rowILClassTarget = wsi[list(wb.defined_names[sNamedRange].destinations)[0][1]].row
        sNamedRange = wsiName + 'TargetDescription'
        rowILDescriptionTarget = wsi[list(wb.defined_names[sNamedRange].destinations)[0][1]].row
    except:
        errorHandler(errProc, errorCode.noNamedRange, sNamedRange)

    #--------------------------------------------------------------------------#
    # Enter a loop to process all of the safety interlocks:                    #
    #--------------------------------------------------------------------------#
    iCol = colILSource
    sNamePrevious = ''
    iRowMax = wsi.max_row
    for i in range(rowILSource + 1, iRowMax + 1):
        #----------------------------------------------------------------------#
        # Check if a new interlock. Exit the loop if none:                     #
        #----------------------------------------------------------------------#
        sName = wsi.cell(row=i, column=colILName).value
        if (sName is None):
            break;
        num = iRowMax + 1 - rowILSource - 1
        pc = pbwt * 1.0 / num
        p.update(pc)
        p.set_description(wsiName + ' ' + sName)
        p.refresh()
        sleep(0.01)
        if (sName != sNamePrevious):
            #------------------------------------------------------------------#
            # Enter a loop to process the interlock name set:                  #
            #------------------------------------------------------------------#
            iRow = i
            iRowNext = iRow
            sExplanation = ''
            sExpression = ''
            sNameNext = wsi.cell(row=iRowNext, column=colILName).value
            while (sNameNext == sName):
                #--------------------------------------------------------------#
                # Get the source device data:                                  #
                #--------------------------------------------------------------#
                sSource = wsi.cell(row=iRowNext, column=colILSource).value
                sCode = wsi.cell(row=iRowNext, column=colILCode).value
                sFunction = wsi.cell(row=iRowNext, column=colILFunction).value
                sFunctionEnd = wsi.cell(row=iRowNext, column=colILFunctionEnd).value
                sState = wsi.cell(row=iRowNext, column=colILState).value
                sDescriptionIL = wsi.cell(row=iRowNext, column=colILDescription).value

                #--------------------------------------------------------------#
                # Replace any placeholder text:                                #
                #--------------------------------------------------------------#
                logging.info('Source: ' + sSource)
                sSource = sSource.replace('@', at)
                getInstanceIDX(sSource)
                sCode = sCode.replace('@@IDX@@', str(idxInstance))
                sCode = sCode.replace('@@CONNECTION@@', str(connInstance))
#                sCode = sCode.replace('@@IDX@@', str(getInstanceIDX(sSource)))
                sCode = sCode.replace('@@STATE@@', sState)
                sDescriptionIL = sDescriptionIL.replace('@', at)
                sDescriptionIL = sDescriptionIL.replace('%', percent)
                sDescriptionIL = sDescriptionIL.replace('$', dollar)
                sDescriptionIL = sSource + ' ' + sDescriptionIL + ' ' + sState

                #--------------------------------------------------------------#
                # Build the interlock code:                                    #
                #--------------------------------------------------------------#
                if (not sFunction is None):
                    sExpression = sExpression + sFunction
                    if (len(sExplanation) == 0):
                        sExplanation = sFunction + ' ' + sDescriptionIL
                    else:
                        sExplanation = sExplanation + '   ' + sFunction + ' ' + sDescriptionIL

                if (sCode[:1] == '('):
                    sExpression = sExpression + sCode
                else:
                    sExpression = sExpression + ' ' + sCode

                if (not sFunctionEnd is None):
                    sExpression = sExpression + '\r\n' + sFunctionEnd

                #--------------------------------------------------------------#
                # Check the next row:                                          #
                #--------------------------------------------------------------#
                iRowNext = iRowNext + 1
                sNameNext = wsi.cell(row=iRowNext, column=colILName).value
                if (sName == sNameNext):
                    sExpression = sExpression + '\r\n'

            #------------------------------------------------------------------#
            # Enter another loop to process all of the safety interlock data   #
            # in the current row for the current interlock source tag:         #
            #------------------------------------------------------------------#
            iXCol = colILTarget + 1
            while (iXCol < colILEnd):
                #--------------------------------------------------------------#
                # Check if an interlock:                                       #
                #--------------------------------------------------------------#
                sX = wsi.cell(row=iRow, column=iXCol).value
                if (sX == sILMarker):
                    #----------------------------------------------------------#
                    # Add the interlock to the output list:                    #
                    #----------------------------------------------------------#
                    sTarget = wsi.cell(row=rowILTarget, column=iXCol).value
                    sTarget = sTarget.replace('@', at)
                    sClassTarget = wsi.cell(row=rowILClassTarget, column=iXCol).value
                    sDescriptionTarget = wsi.cell(row=rowILDescriptionTarget, column=iXCol).value
                    logging.info('Target: ' + sTarget)

                    #----------------------------------------------------------#
                    # Add the interlock to the output list:                    #
                    #----------------------------------------------------------#
                    wso.cell(row=iRowOut, column=1).value = sTarget
                    wso.cell(row=iRowOut, column=2).value = sClassTarget
                    wso.cell(row=iRowOut, column=3).value = str(getInstanceIDX(sTarget))
                    wso.cell(row=iRowOut, column=4).value = sDescriptionTarget
                    wso.cell(row=iRowOut, column=5).value = sExplanation
                    wso.cell(row=iRowOut, column=6).value = sExpression
                    wso.cell(row=iRowOut, column=7).value = connInstance
                    iRowOut = iRowOut + 1

                #--------------------------------------------------------------#
                # Process the next column:                                     #
                #--------------------------------------------------------------#
                iXCol = iXCol + 1

            #------------------------------------------------------------------#
            # Process the next interlock source row:                           #
            #------------------------------------------------------------------#
            sNamePrevious = sName

    #--------------------------------------------------------------------------#
    # Return completion status:                                                #
    #--------------------------------------------------------------------------#
    p.set_description('Processing complete')
    p.refresh()

#------------------------------------------------------------------------------#
# Function getInstanceIDX                                                      #
#                                                                              #
# Description:                                                                 #
# Gets the instance IDX index value from the Instance sheet.                   #
#------------------------------------------------------------------------------#
# Calling parameters:                                                          #
#                                                                              #
# sInstance             The instance name.                                     #
#------------------------------------------------------------------------------#
def getInstanceIDX(sInstance):
    #--------------------------------------------------------------------------#
    # Define the procedure name:                                               #
    #--------------------------------------------------------------------------#
    errProc = getInstanceIDX.__name__

    #--------------------------------------------------------------------------#
    # Define global variables used in this procedure:                          #
    #--------------------------------------------------------------------------#
    global connInstance
    global idxInstance
    global wb

    #--------------------------------------------------------------------------#
    # Get the instance worksheet:                                              #
    #--------------------------------------------------------------------------#
    ws = wb['tblInstance']

    #--------------------------------------------------------------------------#
    # Find the instance name in the list:                                      #
    #--------------------------------------------------------------------------#
    iRow = 2
    while (not ws.cell(row=iRow, column=3).value is None):
        if (ws.cell(row=iRow, column=3).value == sInstance):
            idxInstance = ws.cell(row=iRow, column=7).value
            connInstance = ws.cell(row=iRow, column=21).value
            return;
        else:
            iRow = iRow + 1

    #--------------------------------------------------------------------------#
    # Error if tag not found. Should never get to here:                        #
    #--------------------------------------------------------------------------#
    errorHandler(errProc, errorCode.instanceNotFound, sInstance)

#------------------------------------------------------------------------------#
# Call the main function:                                                      #
#------------------------------------------------------------------------------#
main()
