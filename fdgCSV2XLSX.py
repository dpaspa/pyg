#------------------------------------------------------------------------------#
#            Copyright 2018 Rieckermann Engineering Operations                 #
#------------------------------------------------------------------------------#
# Description:                                                                 #
# This file converts a CSV file to xlsx format.                                #
#------------------------------------------------------------------------------#
# Revision history:                                                            #
# Ver By               Date        CC        Note                              #
# 1   David Paspa      07-Apr-2018 NA        Initial design.                   #
#------------------------------------------------------------------------------#
# Import the python libraries:                                                 #
#------------------------------------------------------------------------------#
from enum import Enum
import os.path
import sys
import csv
import re
import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from tqdm import trange
from time import sleep

import logging
logging.basicConfig(filename='fdg.log',level=logging.DEBUG)

#------------------------------------------------------------------------------#
# Declare global variables:                                                    #
#------------------------------------------------------------------------------#
ps = None

#------------------------------------------------------------------------------#
# Declare the error handling global variables and procedure:                   #
#------------------------------------------------------------------------------#
def errorHandler(errProc, eCode, *args):
    #--------------------------------------------------------------------------#
    # Get the application specific error message and output the error:         #
    #--------------------------------------------------------------------------#
    sMsg = errorMessage[eCode]

    #--------------------------------------------------------------------------#
    # Replace any error parameters:                                            #
    #--------------------------------------------------------------------------#
    i = 1
    for arg in args:
        sMsg = sMsg.replace('@' + str(i), str(arg))
        i = i + 1

    #--------------------------------------------------------------------------#
    # Output the error message and end:                                        #
    #--------------------------------------------------------------------------#
    print('\r\n')
    print('ERROR ' + str(eCode) + ' in Procedure ' + "'" + errProc + "'" + '\r\n' + '\r\n' + sMsg)
    print('\r\n')
    sys.exit()

#------------------------------------------------------------------------------#
# Enumerate the error numbers and map the error messages:                      #
#------------------------------------------------------------------------------#
class errorCode(Enum):
    cannotCreateCSV                    = -1
    cannotOpenCSV                      = -2
    cannotOpenWorkbook                 = -3
    fileNotExist                       = -4
    pathNotExist                       = -5

errorMessage = {
    errorCode.cannotCreateCSV          : 'Cannot create output worksheet @1',
    errorCode.cannotOpenCSV            : 'Cannot open CSV file @1',
    errorCode.cannotOpenWorkbook       : 'Cannot open workbook @1',
    errorCode.fileNotExist             : 'Workbook file @1 does not exist.',
    errorCode.pathNotExist             : 'Ouput path @1 does not exist.'
}

#------------------------------------------------------------------------------#
# Function csv2xlsx                                                            #
#                                                                              #
# Description:                                                                 #
# The CSV to XLSX conversion program.                                          #
#------------------------------------------------------------------------------#
def csv2xlsx(fileCSV, pathOut, fileOutputName):
    #--------------------------------------------------------------------------#
    # Define the procedure name:                                               #
    #--------------------------------------------------------------------------#
    errProc = csv2xlsx.__name__

    #--------------------------------------------------------------------------#
    # Use global variables:                                                    #
    #--------------------------------------------------------------------------#
    global ps

    #--------------------------------------------------------------------------#
    # Create a progress bar:                                                   #
    #--------------------------------------------------------------------------#
    ps = trange(100, desc='Starting...', leave=False)

    #--------------------------------------------------------------------------#
    # Get the input CSV file name and check it exists:                         #
    #--------------------------------------------------------------------------#
    if not os.path.exists(fileCSV):
        errorHandler(errorCode.fileNotExist, fileCSV)

    #--------------------------------------------------------------------------#
    # Open the CSV file:                                                       #
    #--------------------------------------------------------------------------#
#    try:
#        f = open(fileCSV)
#    except:
#        errorHandler(errorCode.cannotOpenCSV, fileCSV)

    #--------------------------------------------------------------------------#
    # Get the base file name and path of the input CSV file:                   #
    #--------------------------------------------------------------------------#
    csvFileName = os.path.basename(fileCSV)
    csvBaseName = os.path.splitext(csvFileName)[0]
    csvDir = os.path.dirname(fileCSV)
    nameSheet = os.path.splitext(csvFileName)[0]

    #--------------------------------------------------------------------------#
    # Get the output path if specified. If not specified then use same path:   #
    #--------------------------------------------------------------------------#
    if (pathOut == ''):
        xlsxDir = csvDir
    else:
        xlsxDir = pathOut
        if not os.path.exists(xlsxDir):
            errorHandler(errProc, errorCode.pathNotExist, xlsxDir)

    #--------------------------------------------------------------------------#
    # Check if the output filename has been specified:                         #
    #--------------------------------------------------------------------------#
    if(len(fileOutputName) > 0):
        xlsxName = xlsxDir + '/' + fileOutputName + '.xlsx'
    else:
        xlsxName = xlsxDir + '/' + csvBaseName + '.xlsx'

    #--------------------------------------------------------------------------#
    # Check if the output spreadsheet already exists. If so add another sheet: #
    #--------------------------------------------------------------------------#
    if not os.path.exists(xlsxName):
        #----------------------------------------------------------------------#
        # Create a new workbook file in memory and use the first worksheet:    #
        #----------------------------------------------------------------------#
        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]
        ws.title = nameSheet
    else:
        wb = openpyxl.load_workbook(filename = xlsxName)
        ws = wb.create_sheet(nameSheet)

    #--------------------------------------------------------------------------#
    # Define the CSV dialect being used:                                       #
    #--------------------------------------------------------------------------#
#    csv.register_dialect('colons', delimiter=':')
#    csv.register_dialect('commas', delimiter=',')
#    reader = csv.reader(f, dialect='commas')

    #--------------------------------------------------------------------------#
    # Get the number of lines in the file:                                     #
    #--------------------------------------------------------------------------#
#    for i, l in enumerate(f):
#        pass
#    numLines = i + 1

    #--------------------------------------------------------------------------#
    # Set the sheet name:                                                      #
    #--------------------------------------------------------------------------#

#wb = openpyxl.Workbook()
#ws = wb.active

    #--------------------------------------------------------------------------#
    # Transfer the CSV data:                                                   #
    #--------------------------------------------------------------------------#
    ps.set_description('Converting: ' + csvBaseName)
    ps.refresh()
    with open(fileCSV) as f:
        reader = csv.reader(f, delimiter=',')
        for row in reader:
            ws.append(row)
#            ws.append([ILLEGAL_CHARACTERS_RE.sub('',row)])

    #--------------------------------------------------------------------------#
    # Transfer the CSV data:                                                   #
    #--------------------------------------------------------------------------#
#    f.seek(0)
#    for row_index, row in enumerate(reader):
#        for column_index, cell in enumerate(row):
#            ws.cell(row=row_index + 1, column=column_index + 1).value = cell
#
#        ps.update(100 * 1.0 / numLines)
#        sleep(0.01)

    #--------------------------------------------------------------------------#
    # Save the changes:                                                        #
    #--------------------------------------------------------------------------#
    wb.save(filename = xlsxName)

    #--------------------------------------------------------------------------#
    # Delete the CSV file:                                                     #
    #--------------------------------------------------------------------------#
#    os.remove(fileCSV)

    #--------------------------------------------------------------------------#
    # Report completion regardless of error:                                   #
    #--------------------------------------------------------------------------#
    ps.set_description(csvBaseName + ': Processing complete')
    ps.refresh()
    ps.close()
    return xlsxName
