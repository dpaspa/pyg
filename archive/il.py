#------------------------------------------------------------------------------#
#            Copyright 2018 Rieckermann Engineering Operations                 #
#------------------------------------------------------------------------------#
# Description:                                                                 #
# This file creates a word report based on the module class.                   #
#------------------------------------------------------------------------------#
# Revision history:                                                            #
# Rev By               Date        CC        Note                              #
# 1.0 David Paspa      05-Apr-2018 NA        Initial design.                   #
#------------------------------------------------------------------------------#
# Import the python libraries:                                                 #
#------------------------------------------------------------------------------#
#from docx import Document
import openpyxl
import sys
import os.path
from enum import Enum
import logging

#------------------------------------------------------------------------------#
# Declare the application title:                                               #
#------------------------------------------------------------------------------#
APP_TITLE = 'Interlock Generator'
APP_VERSION = '1'

#------------------------------------------------------------------------------#
# Declare the error handler:                                                   #
#------------------------------------------------------------------------------#
iErr = 0
errProc = ''
errParameters = []
#    print('Unexpected error:', sys.exc_info()[0])
#    raise

def errorHandler(eCode, *args):
    global iErr
    global errParameters

    iErr = eCode
    errParameters = []
    for arg in args:
        errParameters.append(arg)

#------------------------------------------------------------------------------#
# Map the error number to the error function blocks:                           #
#------------------------------------------------------------------------------#
class errorCode(Enum):
    userCancel = -1
    cannotReplace = -2
    cannotOpenWorkbook = -3
    fileNotExist = -4

errorMessage = {errorCode.userCancel : 'User canceled the operation.',
                errorCode.cannotReplace : 'Cannot replace field @1 with value @2.',
                errorCode.cannotOpenWorkbook : 'Cannot open workbook @1',
                errorCode.fileNotExist : 'Workbook file @1 does not exist.',
}

#errors = ['apple', 'banana', 'grapes', 'pear']
#counter_list = list(enumerate(errors, 1))
#print(counter_list)
# Output: [(1, 'apple'), (2, 'banana'), (3, 'grapes'), (4, 'pear')]

#------------------------------------------------------------------------------#
# Sub generateInterlocks                                                       #
#                                                                              #
# Description:                                                                 #
# Creates tblInterlocks with the interlock information as a list.              #
#------------------------------------------------------------------------------#
# Calling parameters:                                                          #
#                                                                              #
# wb                    The workbook object.                                   #
# sILMarker             The interlock marker, either "X" for critical          #
#                       interlocks or "M" for non-critical interlocks          #
#                       which are allowed to be manually overrridden.          #
# sILSheet              The output sheet name.                                 #
#------------------------------------------------------------------------------#
def generateInterlocks(wb, sILMarker, sILSheet):
    #--------------------------------------------------------------------------#
    # Define the procedure name:                                               #
    #--------------------------------------------------------------------------#
    global errProc
    errProc = generateInterlocks.__name__
#   def divide(x, y):
#       try:
#           result = x / y
#       except ZeroDivisionError:
#           print("division by zero!")
#       else:
#           print("result is", result)
#       finally:
#           print("executing finally clause")

    #wb.get_sheet_names()
    #['Sheet1', 'Sheet2', 'Sheet3']
#    wsi = wb['Sheet1']
    #type(sheet) <class 'openpyxl.worksheet.worksheet.Worksheet'>
#    print(sheet.title)
    #'Sheet3'
    #anotherSheet = wb.active
    #anotherSheet
    #<Worksheet "Sheet1">

    #--------------------------------------------------------------------------#
    # Get the maximum data limits in the input worksheet:                      #
    #--------------------------------------------------------------------------#
#    iMaxRow = sheet.max_row
#    iMaxCol = sheet.max_column

    #--------------------------------------------------------------------------#
    # Get the input worksheet:                                                 #
    #--------------------------------------------------------------------------#
    wsi = wb['tblSafety']

    #--------------------------------------------------------------------------#
    # Delete the output worksheet and create a blank new one:                  #
    #--------------------------------------------------------------------------#
    wso = wb[sILSheet]
    wb.remove(wso)
    wb.create_sheet(sILSheet)
    wso = wb[sILSheet]
    logging.info(wso.title)

    #--------------------------------------------------------------------------#
    # Set the new worksheet titles:                                            #
    #--------------------------------------------------------------------------#
    wso.cell(row=1, column=1).value = 'Source'
    wso.cell(row=1, column=2).value = 'Target'
    wso.cell(row=1, column=3).value = 'Interlock'
    iRowOut = 2

    #--------------------------------------------------------------------------#
    # Get the named ranges to orientate the data:                              #
    #--------------------------------------------------------------------------#
    colILEnd = wsi[list(wb.defined_names['ILEnd'].destinations)[0][1]].col_idx
    colILName = wsi[list(wb.defined_names['ILName'].destinations)[0][1]].col_idx
    colILCode = wsi[list(wb.defined_names['ILCode'].destinations)[0][1]].col_idx
    colILTarget = wsi[list(wb.defined_names['ILTarget'].destinations)[0][1]].col_idx
    rowILTarget = wsi[list(wb.defined_names['ILTarget'].destinations)[0][1]].row
    colILSource = wsi[list(wb.defined_names['ILSource'].destinations)[0][1]].col_idx
    rowILSource = wsi[list(wb.defined_names['ILSource'].destinations)[0][1]].row
    colILState = wsi[list(wb.defined_names['ILState'].destinations)[0][1]].col_idx
    colILFunction = wsi[list(wb.defined_names['ILFunction'].destinations)[0][1]].col_idx
    colILFunctionEnd = wsi[list(wb.defined_names['ILFunctionEnd'].destinations)[0][1]].col_idx
#        print(wsi.cell(row = 6, column = i).value)

    #--------------------------------------------------------------------------#
    # Enter a loop to process all of the safety interlocks:                    #
    #--------------------------------------------------------------------------#
    iCol = colILSource
    sNamePrevious = ''
    for i in range(rowILSource + 1, wsi.max_row):
        #----------------------------------------------------------------------#
        # Check if a new interlock:                                            #
        #----------------------------------------------------------------------#
        sName = wsi.cell(row=i, column=colILName).value
        if (sName != sNamePrevious):
            #------------------------------------------------------------------#
            # Build the interlock code:                                        #
            #------------------------------------------------------------------#
            iRow = i
            iRowNext = i + 1
            sExpression = ''
            sNameNext = wsi.cell(row=iRowNext, column=colILName).value
            while (sNameNext == sName):
                #--------------------------------------------------------------#
                # Get the source device data:                                  #
                #--------------------------------------------------------------#
                sSource = wsi.cell(row=iRowNext, column=colILSource).value
                sCode = wsi.cell(row=iRowNext, column=colILCode).value
                sFunction = wsi.cell(row=iRowNext, column=colILFunction).value
                sFunctionEnd = wsi.cell(row=iRowNext, column=colILFunctionEnd).value
                sState = wsi.cell(row=iRowNext, column=colILState).value

                #--------------------------------------------------------------#
                # Replace any placeholder text:                                #
                #--------------------------------------------------------------#
                logging.info('Source: ' + sSource)
                sCode = sCode.replace('@@INSTANCE@@', sSource)
                sCode = sCode.replace('@@STATE@@', sState)

                #--------------------------------------------------------------#
                # Build the interlock code:                                    #
                #--------------------------------------------------------------#
                if (not sFunction is None):
                    sExpression = sExpression + sFunction

                sExpression = sExpression + '\r\n' + sCode

                if (not sFunctionEnd is None):
                    sExpression = sExpression + '\r\n' + sFunctionEnd

                #--------------------------------------------------------------#
                # Check the next row:                                          #
                #--------------------------------------------------------------#
                iRowNext = iRowNext + 1
                sNameNext = wsi.cell(row=iRowNext, column=colILName).value
                if (sName == sNameNext):
                    sExpression = sExpression + '\r\n'

            #------------------------------------------------------------------#
            # Enter another loop to process all of the safety interlock data   #
            # in the current row for the current interlock source tag:         #
            #------------------------------------------------------------------#
            iXCol = colILTarget + 1
            while (iXCol < colILEnd):
                #--------------------------------------------------------------#
                # Check if an interlock:                                       #
                #--------------------------------------------------------------#
                sX = wsi.cell(row=iRow, column=iXCol).value
                if (sX == sILMarker):
                    #----------------------------------------------------------#
                    # Add the interlock to the output list:                    #
                    #----------------------------------------------------------#
                    sTarget = wsi.cell(row=rowILTarget, column=iXCol).value
                    logging.info('Target: ' + sTarget)

                    #----------------------------------------------------------#
                    # Add the interlock to the output list:                    #
                    #----------------------------------------------------------#
                    wso.cell(row=iRowOut, column=1).value = sSource
                    wso.cell(row=iRowOut, column=2).value = sTarget
                    wso.cell(row=iRowOut, column=3).value = sExpression
                    iRowOut = iRowOut + 1

                #--------------------------------------------------------------#
                # Process the next column:                                     #
                #--------------------------------------------------------------#
                iXCol = iXCol + 1

            #------------------------------------------------------------------#
            # Process the next interlock source row:                           #
            #------------------------------------------------------------------#
            sNamePrevious = sName

    #--------------------------------------------------------------------------#
    # Return completion status:                                                #
    #--------------------------------------------------------------------------#
    return iErr

#------------------------------------------------------------------------------#
# Sub AppOutputStatus                                                          #
#                                                                              #
# Description:                                                                 #
# Gets a handle to Word and opens the specified document file or else          #
# just creates a new blank one.                                                #
#------------------------------------------------------------------------------#
# Calling parameters:                                                          #
#                                                                              #
# none                                                                         #
#------------------------------------------------------------------------------#
# Returned parameters:                                                         #
#                                                                              #
# none                                                                         #
#------------------------------------------------------------------------------#
def reportComplete():
    #--------------------------------------------------------------------------#
    # Declare global parameters:                                               #
    #--------------------------------------------------------------------------#
    global iErr
    global errProc
    global errParameters

    #--------------------------------------------------------------------------#
    # Check if successful completion:                                          #
    #--------------------------------------------------------------------------#
    if (iErr == 0):
        #----------------------------------------------------------------------#
        # Output a success message:                                            #
        #----------------------------------------------------------------------#
        print('Congratulations! Operation successful.')
    else:
        #----------------------------------------------------------------------#
        # Get the application specific error message and output the error:     #
        #----------------------------------------------------------------------#
        sMsg = errorMessage[iErr]

        #----------------------------------------------------------------------#
        # Check if there are any error parameters to add:                      #
        #----------------------------------------------------------------------#
        if (not errParameters is None):
            #------------------------------------------------------------------#
            # Enter a loop to add each parameter:                              #
            #------------------------------------------------------------------#
            for i in range(0, len(errParameters)):
                sMsg = sMsg.replace('@' + str(i + 1), errParameters[i])

        #----------------------------------------------------------------------#
        # Output the error message:                                            #
        #----------------------------------------------------------------------#
        logging.critical(APP_TITLE + ' Version ' + APP_VERSION + '\r\n' + 'ERROR ' + str(iErr) + ' in Procedure ' + "'" + errProc + "'" + '\r\n' + sMsg)

#my_file = Path("/path/to/file")
#if my_file.is_file():
#    print sys.argv[0] # prints python_script.py
#    print sys.argv[1] # prints var1
#    print sys.argv[2] # prints var2
#------------------------------------------------------------------------------#
# Define the main functions:                                                   #
#------------------------------------------------------------------------------#
def main():
    #--------------------------------------------------------------------------#
    # Define the procedure name:                                               #
    #--------------------------------------------------------------------------#
    global errProc
    errProc = main.__name__
    logging.basicConfig(level=logging.INFO)

    #--------------------------------------------------------------------------#
    # Get the interlock configuration workbook name and check it exists:       #
    #--------------------------------------------------------------------------#
    wbName = sys.argv[1]
    if not os.path.exists(wbName):
        errorHandler(errorCode.fileNotExist, wbName)
    else:
        #----------------------------------------------------------------------#
        # Open the workbook:                                                   #
        #----------------------------------------------------------------------#
        try:
            wb = openpyxl.load_workbook(wbName, data_only=True)

        except:
            errorHandler(errorCode.cannotOpenWorkbook, wbName)
        else:
            #------------------------------------------------------------------#
            # Process the critical interlocks:                                 #
            #------------------------------------------------------------------#
            if (generateInterlocks(wb, 'X', 'tblInterlockCR') != 0):
                pass

            #------------------------------------------------------------------#
            # Process the non-critical interlocks:                             #
            #------------------------------------------------------------------#
            elif (generateInterlocks(wb, 'M', 'tblInterlockNCR') != 0):
                pass

    #--------------------------------------------------------------------------#
    # Save the changes if no error:                                            #
    #--------------------------------------------------------------------------#
    if (iErr == 0):
        wb.save('il.xlsx')

    reportComplete()

#------------------------------------------------------------------------------#
# Call the main function:                                                      #
#------------------------------------------------------------------------------#
main()
